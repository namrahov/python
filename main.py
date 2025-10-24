# def greet(name: str) -> str:
#     return f"Hello, {name}!"
#
# if __name__ == "__main__":
#     print(greet("World"))
import os
import openpyxl
import shutil

from service.UserService import save_all
from db.Database import Base, engine

Base.metadata.create_all(bind=engine)

if __name__ == "__main__":

    os.chdir('C:\\Users\\AZINTELECOM\\OneDrive\\İş masası')
    workbook = openpyxl.load_workbook('siyahi.xlsx', read_only = True)
    users = []

    for sheet in workbook.sheetnames:
        ws = workbook[sheet]

    for row in ws.iter_rows(values_only=True):
        print(row)
        if row is not None:
            name = row[0]
            email = row[1]
            users.append({"name": name, "email": email})

    if users:
        save_all(users)
    else:
        print("⚠️ No valid rows found to save.")


   # save_user("Tarkhan", "tarkhan@example.com")
   #shutil.copy('C:\\Users\\AZINTELECOM\\IdeaProjects\\spam.txt', 'C:\\Users\\AZINTELECOM')
 # for folderName, subfolders, filenames in os.walk('C:\\Users\\AZINTELECOM\\IdeaProjects'):
 #       print(filenames)
       #print(subfolders)
       #print(folderName)

